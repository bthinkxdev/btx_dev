"""
Finance management reports — aggregated queries only (no statutory accounting).

Designed for large tables: date-range filters first, values()+annotate aggregations,
iterator() for row exports. Avoids N+1 via select_related on detail lists.
"""

from __future__ import annotations

import csv
import io
import json
from calendar import monthrange
from dataclasses import dataclass, field
from datetime import date, timedelta
from decimal import Decimal
from typing import Any, Iterable

from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncMonth, TruncWeek, TruncYear
from django.http import HttpResponse
from django.utils import timezone

from ..models import (
    Client,
    Expense,
    ExpenseCategory,
    Income,
    IncomeCategory,
    Project,
)

REPORT_TYPES = {
    'income': 'Income Report',
    'expense': 'Expense Report',
    'profit': 'Profit Report',
    'cash_flow': 'Cash Flow',
    'project_profitability': 'Project Profitability',
    'client_revenue': 'Client Revenue',
    'expense_category': 'Expense Category Report',
    'payment_method': 'Payment Method Report',
    'top_customers': 'Top Customers',
    'top_projects': 'Top Projects',
}


def _money(value) -> Decimal:
    if value is None:
        return Decimal('0')
    return Decimal(value)


def _growth_pct(current: Decimal, previous: Decimal) -> float:
    if previous == 0:
        return 100.0 if current else 0.0
    return float(((current - previous) / previous) * Decimal('100'))


@dataclass
class ReportFilters:
    period: str = 'month'  # day|week|month|year|custom
    date_from: date | None = None
    date_to: date | None = None
    project_id: int | None = None
    client_id: int | None = None
    category_id: int | None = None  # income or expense category depending on report
    employee_id: int | None = None
    payment_method: str = ''
    payment_status: str = ''
    payment_type: str = ''  # income payment type

    start: date = field(default_factory=timezone.localdate)
    end: date = field(default_factory=timezone.localdate)
    label: str = 'This Month'


def resolve_report_period(
    period: str,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    ref: date | None = None,
) -> tuple[date, date, str]:
    today = ref or timezone.localdate()
    period = (period or 'month').strip().lower()
    if period in ('day', 'daily', 'today'):
        return today, today, 'Daily'
    if period in ('week', 'weekly'):
        start = today - timedelta(days=today.weekday())
        return start, start + timedelta(days=6), 'Weekly'
    if period in ('year', 'yearly'):
        return date(today.year, 1, 1), date(today.year, 12, 31), 'Yearly'
    if period == 'custom' and date_from and date_to:
        if date_from > date_to:
            date_from, date_to = date_to, date_from
        return date_from, date_to, 'Custom'
    # monthly default
    start = today.replace(day=1)
    end = today.replace(day=monthrange(today.year, today.month)[1])
    return start, end, 'Monthly'


def filters_from_cleaned(data: dict) -> ReportFilters:
    period = data.get('period') or 'month'
    start, end, label = resolve_report_period(
        period,
        date_from=data.get('date_from'),
        date_to=data.get('date_to'),
    )
    project = data.get('project')
    client = data.get('client')
    category = data.get('category')
    employee = data.get('employee')
    return ReportFilters(
        period=period,
        date_from=data.get('date_from'),
        date_to=data.get('date_to'),
        project_id=getattr(project, 'pk', project) or None,
        client_id=getattr(client, 'pk', client) or None,
        category_id=getattr(category, 'pk', category) or None,
        employee_id=getattr(employee, 'pk', employee) or None,
        payment_method=(data.get('payment_method') or '').strip(),
        payment_status=(data.get('payment_status') or '').strip(),
        payment_type=(data.get('payment_type') or '').strip(),
        start=start,
        end=end,
        label=label,
    )


def _income_qs(f: ReportFilters):
    qs = Income.objects.filter(payment_date__gte=f.start, payment_date__lte=f.end)
    if f.project_id:
        qs = qs.filter(project_id=f.project_id)
    if f.client_id:
        qs = qs.filter(client_id=f.client_id)
    if f.category_id:
        qs = qs.filter(category_id=f.category_id)
    if f.payment_type:
        qs = qs.filter(payment_type=f.payment_type)
    if f.payment_status:
        qs = qs.filter(payment_status=f.payment_status)
    if f.payment_method:
        # Income uses payment_type; map method filter when provided
        qs = qs.filter(payment_type=f.payment_method)
    return qs


def _expense_qs(f: ReportFilters):
    qs = Expense.objects.filter(expense_date__gte=f.start, expense_date__lte=f.end)
    if f.project_id:
        qs = qs.filter(project_id=f.project_id)
    if f.client_id:
        qs = qs.filter(project__client_id=f.client_id)
    if f.category_id:
        qs = qs.filter(category_id=f.category_id)
    if f.employee_id:
        qs = qs.filter(employee_id=f.employee_id)
    if f.payment_method:
        qs = qs.filter(payment_method=f.payment_method)
    return qs


def _sum(qs) -> Decimal:
    return _money(qs.aggregate(t=Sum('amount'))['t'])


def _trunc_for_period(period: str):
    """Return a Trunc* function, or None to group by the raw DateField.

    TruncDate on DateField raises OperationalError on SQLite with USE_TZ.
    """
    p = (period or 'month').lower()
    if p in ('day', 'daily', 'today', 'custom'):
        return None
    if p in ('week', 'weekly'):
        return TruncWeek
    if p in ('year', 'yearly'):
        return TruncYear
    return TruncMonth


def _series_labels(buckets: list, period: str) -> list[str]:
    out = []
    for b in buckets:
        if b is None:
            out.append('—')
            continue
        d = b.date() if hasattr(b, 'date') and not isinstance(b, date) else b
        p = (period or 'month').lower()
        if p in ('day', 'daily', 'today', 'custom'):
            out.append(d.strftime('%d %b'))
        elif p in ('week', 'weekly'):
            out.append(d.strftime('W%W %Y'))
        elif p in ('year', 'yearly'):
            out.append(d.strftime('%Y'))
        else:
            out.append(d.strftime('%b %Y'))
    return out


def build_time_series(f: ReportFilters) -> dict[str, Any]:
    Trunc = _trunc_for_period(f.period if f.period != 'custom' else 'day')
    if Trunc is None:
        income_rows = list(
            _income_qs(f)
            .values('payment_date')
            .annotate(total=Sum('amount'))
            .order_by('payment_date')
        )
        expense_rows = list(
            _expense_qs(f)
            .values('expense_date')
            .annotate(total=Sum('amount'))
            .order_by('expense_date')
        )
        imap = {r['payment_date']: _money(r['total']) for r in income_rows if r['payment_date']}
        emap = {r['expense_date']: _money(r['total']) for r in expense_rows if r['expense_date']}
    else:
        income_rows = list(
            _income_qs(f)
            .annotate(bucket=Trunc('payment_date'))
            .values('bucket')
            .annotate(total=Sum('amount'))
            .order_by('bucket')
        )
        expense_rows = list(
            _expense_qs(f)
            .annotate(bucket=Trunc('expense_date'))
            .values('bucket')
            .annotate(total=Sum('amount'))
            .order_by('bucket')
        )
        imap = {r['bucket']: _money(r['total']) for r in income_rows if r['bucket']}
        emap = {r['bucket']: _money(r['total']) for r in expense_rows if r['bucket']}

    keys = sorted(set(imap) | set(emap), key=lambda x: x)
    rev = [float(imap.get(k, 0)) for k in keys]
    exp = [float(emap.get(k, 0)) for k in keys]
    profit = [round(rev[i] - exp[i], 2) for i in range(len(keys))]
    return {
        'labels': _series_labels(keys, f.period),
        'revenue': rev,
        'expense': exp,
        'profit': profit,
        'chart_labels': json.dumps(_series_labels(keys, f.period)),
        'chart_revenue': json.dumps(rev),
        'chart_expense': json.dumps(exp),
        'chart_profit': json.dumps(profit),
    }


def run_report(report_type: str, f: ReportFilters) -> dict[str, Any]:
    """Return context dict for a named management report."""
    report_type = (report_type or '').strip().lower()
    title = REPORT_TYPES.get(report_type, 'Finance Report')
    income_total = _sum(_income_qs(f))
    expense_total = _sum(_expense_qs(f))
    profit_total = income_total - expense_total
    series = build_time_series(f)

    base = {
        'report_type': report_type,
        'report_title': title,
        'filters': f,
        'period_label': f.label,
        'filter_start': f.start,
        'filter_end': f.end,
        'income_total': income_total,
        'expense_total': expense_total,
        'profit_total': profit_total,
        **series,
        'columns': [],
        'rows': [],
        'summary': {},
        'breakdown_labels': '[]',
        'breakdown_values': '[]',
        'detail_note': '',
    }

    builders = {
        'income': _report_income,
        'expense': _report_expense,
        'profit': _report_profit,
        'cash_flow': _report_cash_flow,
        'project_profitability': _report_project_profitability,
        'client_revenue': _report_client_revenue,
        'expense_category': _report_expense_category,
        'payment_method': _report_payment_method,
        'top_customers': _report_top_customers,
        'top_projects': _report_top_projects,
    }
    builder = builders.get(report_type, _report_profit)
    extra = builder(f, income_total, expense_total, profit_total)
    base.update(extra)
    return base


def _report_income(f, income_total, expense_total, profit_total) -> dict:
    rows = list(
        _income_qs(f)
        .values('payment_date', 'category__name', 'client__business_name', 'project_id',
                'payment_type', 'payment_status', 'reference', 'amount')
        .order_by('-payment_date', '-id')[:500]
    )
    by_cat = list(
        _income_qs(f)
        .values('category__name')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    return {
        'columns': ['Date', 'Category', 'Client', 'Project', 'Type', 'Status', 'Reference', 'Amount'],
        'rows': [
            [
                r['payment_date'],
                r['category__name'] or '—',
                r['client__business_name'] or '—',
                f"#{r['project_id']}" if r['project_id'] else '—',
                r['payment_type'],
                r['payment_status'],
                r['reference'] or '—',
                r['amount'],
            ]
            for r in rows
        ],
        'summary': {'Total Income': income_total, 'Entries': len(rows)},
        'breakdown_labels': json.dumps([x['category__name'] or '—' for x in by_cat]),
        'breakdown_values': json.dumps([float(_money(x['total'])) for x in by_cat]),
        'detail_note': 'Showing latest 500 matching income rows (aggregates use full range).',
    }


def _report_expense(f, income_total, expense_total, profit_total) -> dict:
    rows = list(
        _expense_qs(f)
        .values(
            'expense_date', 'category__name', 'vendor', 'project_id',
            'employee__username', 'payment_method', 'amount',
        )
        .order_by('-expense_date', '-id')[:500]
    )
    by_cat = list(
        _expense_qs(f)
        .values('category__name')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    return {
        'columns': ['Date', 'Category', 'Vendor', 'Project', 'Employee', 'Method', 'Amount'],
        'rows': [
            [
                r['expense_date'],
                r['category__name'] or '—',
                r['vendor'] or '—',
                f"#{r['project_id']}" if r['project_id'] else '—',
                r['employee__username'] or '—',
                r['payment_method'],
                r['amount'],
            ]
            for r in rows
        ],
        'summary': {'Total Expense': expense_total, 'Entries': len(rows)},
        'breakdown_labels': json.dumps([x['category__name'] or '—' for x in by_cat]),
        'breakdown_values': json.dumps([float(_money(x['total'])) for x in by_cat]),
        'detail_note': 'Showing latest 500 matching expense rows (aggregates use full range).',
    }


def _report_profit(f, income_total, expense_total, profit_total) -> dict:
    series = build_time_series(f)
    rows = [
        [series['labels'][i], series['revenue'][i], series['expense'][i], series['profit'][i]]
        for i in range(len(series['labels']))
    ]
    return {
        'columns': ['Period', 'Revenue', 'Expense', 'Profit'],
        'rows': rows,
        'summary': {
            'Revenue': income_total,
            'Expense': expense_total,
            'Profit': profit_total,
        },
        'breakdown_labels': json.dumps(['Revenue', 'Expense', 'Profit']),
        'breakdown_values': json.dumps([
            float(income_total), float(expense_total), float(profit_total),
        ]),
    }


def _report_cash_flow(f, income_total, expense_total, profit_total) -> dict:
    """Cash in/out by payment method (management view, not bank reconciliation)."""
    in_by = list(
        _income_qs(f)
        .values('payment_type')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    out_by = list(
        _expense_qs(f)
        .values('payment_method')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    type_labels = dict(Income.PaymentType.choices)
    method_labels = dict(Expense.PaymentMethod.choices)
    rows = []
    for r in in_by:
        rows.append([
            'In',
            type_labels.get(r['payment_type'], r['payment_type']),
            r['count'],
            r['total'],
        ])
    for r in out_by:
        rows.append([
            'Out',
            method_labels.get(r['payment_method'], r['payment_method']),
            r['count'],
            r['total'],
        ])
    return {
        'columns': ['Direction', 'Method', 'Count', 'Amount'],
        'rows': rows,
        'summary': {
            'Cash In (all methods)': income_total,
            'Cash Out (all methods)': expense_total,
            'Net': profit_total,
        },
        'breakdown_labels': json.dumps(['In', 'Out']),
        'breakdown_values': json.dumps([float(income_total), float(expense_total)]),
    }


def _report_project_profitability(f, income_total, expense_total, profit_total) -> dict:
    # Aggregate income & expense per project in two queries, merge in Python
    inc = {
        r['project_id']: _money(r['total'])
        for r in (
            _income_qs(f)
            .filter(project_id__isnull=False)
            .values('project_id')
            .annotate(total=Sum('amount'))
        )
    }
    exp = {
        r['project_id']: _money(r['total'])
        for r in (
            _expense_qs(f)
            .filter(project_id__isnull=False)
            .values('project_id')
            .annotate(total=Sum('amount'))
        )
    }
    pids = set(inc) | set(exp)
    projects = {
        p.pk: p
        for p in Project.objects.filter(pk__in=pids)
        .select_related('client', 'package')
        .only('id', 'client__business_name', 'package__name', 'deal_value', 'balance_due')
    }
    rows = []
    chart_labels, chart_values = [], []
    for pid in sorted(pids, key=lambda i: float(inc.get(i, 0) - exp.get(i, 0)), reverse=True):
        p = projects.get(pid)
        rev = inc.get(pid, Decimal('0'))
        cost = exp.get(pid, Decimal('0'))
        profit = rev - cost
        margin = float((profit / rev) * 100) if rev else 0.0
        name = (
            f"#{pid} {p.client.business_name}" if p else f"#{pid}"
        )
        rows.append([
            name,
            p.deal_value if p else '—',
            rev,
            cost,
            profit,
            f'{margin:.1f}%',
            p.balance_due if p else '—',
        ])
        chart_labels.append(name[:28])
        chart_values.append(float(profit))
    return {
        'columns': ['Project', 'Deal', 'Revenue', 'Expense', 'Profit', 'Margin', 'Pending'],
        'rows': rows[:200],
        'summary': {'Projects': len(rows), 'Net Profit': profit_total},
        'breakdown_labels': json.dumps(chart_labels[:12]),
        'breakdown_values': json.dumps(chart_values[:12]),
    }


def _report_client_revenue(f, income_total, expense_total, profit_total) -> dict:
    rows_raw = list(
        _income_qs(f)
        .filter(client_id__isnull=False)
        .values('client_id', 'client__business_name')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')[:200]
    )
    return {
        'columns': ['Client', 'Payments', 'Revenue'],
        'rows': [
            [r['client__business_name'], r['count'], r['total']]
            for r in rows_raw
        ],
        'summary': {'Clients': len(rows_raw), 'Total Revenue': income_total},
        'breakdown_labels': json.dumps([r['client__business_name'] for r in rows_raw[:12]]),
        'breakdown_values': json.dumps([float(_money(r['total'])) for r in rows_raw[:12]]),
    }


def _report_expense_category(f, income_total, expense_total, profit_total) -> dict:
    rows_raw = list(
        _expense_qs(f)
        .values('category__name')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    return {
        'columns': ['Category', 'Count', 'Amount'],
        'rows': [[r['category__name'], r['count'], r['total']] for r in rows_raw],
        'summary': {'Total Expense': expense_total},
        'breakdown_labels': json.dumps([r['category__name'] or '—' for r in rows_raw]),
        'breakdown_values': json.dumps([float(_money(r['total'])) for r in rows_raw]),
    }


def _report_payment_method(f, income_total, expense_total, profit_total) -> dict:
    income_rows = list(
        _income_qs(f)
        .values('payment_type')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    expense_rows = list(
        _expense_qs(f)
        .values('payment_method')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    type_labels = dict(Income.PaymentType.choices)
    method_labels = dict(Expense.PaymentMethod.choices)
    rows = []
    for r in income_rows:
        rows.append([
            'Income',
            type_labels.get(r['payment_type'], r['payment_type']),
            r['count'],
            r['total'],
        ])
    for r in expense_rows:
        rows.append([
            'Expense',
            method_labels.get(r['payment_method'], r['payment_method']),
            r['count'],
            r['total'],
        ])
    return {
        'columns': ['Source', 'Method', 'Count', 'Amount'],
        'rows': rows,
        'summary': {'Income': income_total, 'Expense': expense_total},
        'breakdown_labels': json.dumps([
            f"In:{type_labels.get(r['payment_type'], r['payment_type'])}" for r in income_rows
        ] + [
            f"Out:{method_labels.get(r['payment_method'], r['payment_method'])}"
            for r in expense_rows
        ]),
        'breakdown_values': json.dumps(
            [float(_money(r['total'])) for r in income_rows]
            + [float(_money(r['total'])) for r in expense_rows]
        ),
    }


def _report_top_customers(f, income_total, expense_total, profit_total) -> dict:
    return _report_client_revenue(f, income_total, expense_total, profit_total)


def _report_top_projects(f, income_total, expense_total, profit_total) -> dict:
    rows_raw = list(
        _income_qs(f)
        .filter(project_id__isnull=False)
        .values('project_id', 'project__client__business_name')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')[:50]
    )
    return {
        'columns': ['Project', 'Client', 'Payments', 'Revenue'],
        'rows': [
            [
                f"#{r['project_id']}",
                r['project__client__business_name'] or '—',
                r['count'],
                r['total'],
            ]
            for r in rows_raw
        ],
        'summary': {'Projects': len(rows_raw), 'Total Revenue': income_total},
        'breakdown_labels': json.dumps([
            f"#{r['project_id']} {r['project__client__business_name'] or ''}"[:28]
            for r in rows_raw[:12]
        ]),
        'breakdown_values': json.dumps([float(_money(r['total'])) for r in rows_raw[:12]]),
    }


def get_executive_dashboard(*, ref: date | None = None) -> dict[str, Any]:
    """Current-month executive KPIs + growth vs previous month."""
    today = ref or timezone.localdate()
    start = today.replace(day=1)
    end = today.replace(day=monthrange(today.year, today.month)[1])
    prev_end = start - timedelta(days=1)
    prev_start = prev_end.replace(day=1)

    def rng(a, b):
        return ReportFilters(start=a, end=b, period='month', label='Monthly')

    cur = rng(start, end)
    prev = rng(prev_start, prev_end)

    revenue = _sum(_income_qs(cur))
    expense = _sum(_expense_qs(cur))
    profit = revenue - expense
    prev_rev = _sum(_income_qs(prev))
    prev_exp = _sum(_expense_qs(prev))

    collections = revenue  # received this month
    pending = _money(
        Project.objects.filter(balance_due__gt=0).aggregate(t=Sum('balance_due'))['t']
    )
    days = max(today.day, 1)
    avg_rev = (revenue / days).quantize(Decimal('0.01'))
    avg_exp = (expense / days).quantize(Decimal('0.01'))

    series = build_time_series(cur)
    top_clients = _report_top_customers(cur, revenue, expense, profit)
    top_projects = _report_top_projects(cur, revenue, expense, profit)

    return {
        'today': today,
        'month_start': start,
        'month_end': end,
        'revenue': revenue,
        'expense': expense,
        'profit': profit,
        'collections': collections,
        'pending': pending,
        'growth': _growth_pct(revenue, prev_rev),
        'expense_growth': _growth_pct(expense, prev_exp),
        'avg_revenue': avg_rev,
        'avg_expense': avg_exp,
        'prev_revenue': prev_rev,
        'prev_expense': prev_exp,
        'labels': series['labels'],
        'chart_labels': series['chart_labels'],
        'chart_revenue': series['chart_revenue'],
        'chart_expense': series['chart_expense'],
        'chart_profit': series['chart_profit'],
        'top_customers_rows': top_clients['rows'][:8],
        'top_projects_rows': top_projects['rows'][:8],
        'breakdown_labels': top_clients.get('breakdown_labels', '[]'),
        'breakdown_values': top_clients.get('breakdown_values', '[]'),
    }


# ── Exports ───────────────────────────────────────────────────────────────────

def _cell(value) -> str:
    if value is None:
        return ''
    if isinstance(value, Decimal):
        return f'{value:.2f}'
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def export_csv(report: dict) -> HttpResponse:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([report['report_title'], report['period_label'],
                     f"{report['filter_start']} – {report['filter_end']}"])
    writer.writerow([])
    for k, v in (report.get('summary') or {}).items():
        writer.writerow([k, _cell(v)])
    writer.writerow([])
    writer.writerow(report.get('columns') or [])
    for row in report.get('rows') or []:
        writer.writerow([_cell(c) for c in row])
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    fname = f"{report['report_type']}_{report['filter_start']}_{report['filter_end']}.csv"
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_xlsx(report: dict) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (report['report_title'] or 'Report')[:31]
    ws.append([report['report_title']])
    ws.append([f"{report['period_label']}: {report['filter_start']} – {report['filter_end']}"])
    ws.append([])
    for k, v in (report.get('summary') or {}).items():
        ws.append([k, _cell(v)])
    ws.append([])
    cols = report.get('columns') or []
    ws.append(cols)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for row in report.get('rows') or []:
        ws.append([_cell(c) for c in row])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"{report['report_type']}_{report['filter_start']}_{report['filter_end']}.xlsx"
    resp = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_pdf(report: dict) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(report['report_title'], styles['Title']),
        Paragraph(
            f"{report['period_label']} · {report['filter_start']} – {report['filter_end']}",
            styles['Normal'],
        ),
        Spacer(1, 8),
    ]
    for k, v in (report.get('summary') or {}).items():
        story.append(Paragraph(f"<b>{k}:</b> {_cell(v)}", styles['Normal']))
    story.append(Spacer(1, 10))

    cols = report.get('columns') or []
    data = [cols] + [
        [_cell(c) for c in row] for row in (report.get('rows') or [])[:200]
    ]
    if data and data[0]:
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(table)
    doc.build(story)
    buf.seek(0)
    fname = f"{report['report_type']}_{report['filter_start']}_{report['filter_end']}.pdf"
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
