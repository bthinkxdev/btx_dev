"""
Finance Statement — chronological Income + Expense with running balance.

Management cashbook view (not statutory accounting). Opening balance is the
net of all matching transactions before the selected range; closing is the
available business balance after the range.
"""

from __future__ import annotations

import calendar
import csv
import io
from dataclasses import dataclass, field
from datetime import date, timedelta
from decimal import Decimal
from typing import Any

from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.utils import timezone

from ..models import Expense, Income

STATEMENT_PER_PAGE = 50
STATEMENT_EXPORT_ROW_CAP = 20_000


def _money(value) -> Decimal:
    if value is None:
        return Decimal('0')
    return Decimal(value)


def resolve_statement_period(
    period: str,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    ref: date | None = None,
) -> tuple[date, date, str]:
    today = ref or timezone.localdate()
    period = (period or 'month').strip().lower()
    if period == 'today':
        return today, today, 'Today'
    if period == 'yesterday':
        y = today - timedelta(days=1)
        return y, y, 'Yesterday'
    if period in ('week', 'this_week'):
        start = today - timedelta(days=today.weekday())
        return start, start + timedelta(days=6), 'This Week'
    if period in ('last_month', 'prev_month'):
        first = today.replace(day=1)
        last_prev = first - timedelta(days=1)
        start = last_prev.replace(day=1)
        return start, last_prev, 'Last Month'
    if period == 'custom' and date_from and date_to:
        if date_from > date_to:
            date_from, date_to = date_to, date_from
        return date_from, date_to, 'Custom Range'
    start = today.replace(day=1)
    end = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    return start, end, 'This Month'


@dataclass
class StatementFilters:
    period: str = 'month'
    date_from: date | None = None
    date_to: date | None = None
    project_id: int | None = None
    client_id: int | None = None
    income_category_id: int | None = None
    expense_category_id: int | None = None
    q: str = ''
    sort: str = 'newest'  # newest | oldest | amount
    start: date = field(default_factory=timezone.localdate)
    end: date = field(default_factory=timezone.localdate)
    label: str = 'This Month'

def filters_from_cleaned(data: dict) -> StatementFilters:
    period = data.get('period') or 'month'
    start, end, label = resolve_statement_period(
        period,
        date_from=data.get('date_from'),
        date_to=data.get('date_to'),
    )
    project = data.get('project')
    client = data.get('client')
    inc_cat = data.get('income_category')
    exp_cat = data.get('expense_category')
    return StatementFilters(
        period=period,
        date_from=data.get('date_from'),
        date_to=data.get('date_to'),
        project_id=getattr(project, 'pk', project) or None,
        client_id=getattr(client, 'pk', client) or None,
        income_category_id=getattr(inc_cat, 'pk', inc_cat) or None,
        expense_category_id=getattr(exp_cat, 'pk', exp_cat) or None,
        q=(data.get('q') or '').strip(),
        sort=(data.get('sort') or 'newest').strip().lower(),
        start=start,
        end=end,
        label=label,
    )


def _apply_income_dims(qs, f: StatementFilters, *, before: bool = False, in_range: bool = False):
    if before:
        qs = qs.filter(payment_date__lt=f.start)
    elif in_range:
        qs = qs.filter(payment_date__gte=f.start, payment_date__lte=f.end)
    if f.project_id:
        qs = qs.filter(project_id=f.project_id)
    if f.client_id:
        qs = qs.filter(client_id=f.client_id)
    if f.income_category_id:
        qs = qs.filter(category_id=f.income_category_id)
    return qs


def _apply_expense_dims(qs, f: StatementFilters, *, before: bool = False, in_range: bool = False):
    if before:
        qs = qs.filter(expense_date__lt=f.start)
    elif in_range:
        qs = qs.filter(expense_date__gte=f.start, expense_date__lte=f.end)
    if f.project_id:
        qs = qs.filter(project_id=f.project_id)
    if f.client_id:
        qs = qs.filter(project__client_id=f.client_id)
    if f.expense_category_id:
        qs = qs.filter(category_id=f.expense_category_id)
    return qs


def _apply_income_search(qs, q: str):
    if not q:
        return qs
    return qs.filter(
        Q(reference__icontains=q)
        | Q(notes__icontains=q)
        | Q(bank_account__icontains=q)
        | Q(client__business_name__icontains=q)
        | Q(category__name__icontains=q)
        | Q(project__client__business_name__icontains=q)
    )


def _apply_expense_search(qs, q: str):
    if not q:
        return qs
    return qs.filter(
        Q(vendor__icontains=q)
        | Q(notes__icontains=q)
        | Q(paid_from__icontains=q)
        | Q(category__name__icontains=q)
        | Q(project__client__business_name__icontains=q)
        | Q(employee__username__icontains=q)
        | Q(employee__first_name__icontains=q)
    )


def opening_balance(f: StatementFilters) -> Decimal:
    """Net of all matching transactions strictly before the range start."""
    inc = _money(
        _apply_income_dims(Income.objects.all(), f, before=True).aggregate(t=Sum('amount'))['t']
    )
    exp = _money(
        _apply_expense_dims(Expense.objects.all(), f, before=True).aggregate(t=Sum('amount'))['t']
    )
    return (inc - exp).quantize(Decimal('0.01'))


def period_totals(f: StatementFilters) -> tuple[Decimal, Decimal]:
    """Period income/expense totals (ignores search so KPIs stay stable with search)."""
    inc_qs = _apply_income_dims(Income.objects.all(), f, in_range=True)
    exp_qs = _apply_expense_dims(Expense.objects.all(), f, in_range=True)
    # Search applies to statement rows; KPIs use dimensional filters only
    income = _money(inc_qs.aggregate(t=Sum('amount'))['t'])
    expense = _money(exp_qs.aggregate(t=Sum('amount'))['t'])
    return income, expense


def _created_by_label(row: dict) -> str:
    first = (row.get('created_by__first_name') or '').strip()
    last = (row.get('created_by__last_name') or '').strip()
    full = f'{first} {last}'.strip()
    if full:
        return full
    return row.get('created_by__username') or '—'


def _fetch_period_rows(f: StatementFilters, *, apply_search: bool = False) -> list[dict[str, Any]]:
    """Lightweight dict rows for the period."""
    inc_qs = _apply_income_dims(Income.objects.all(), f, in_range=True)
    exp_qs = _apply_expense_dims(Expense.objects.all(), f, in_range=True)
    if apply_search and f.q:
        inc_qs = _apply_income_search(inc_qs, f.q)
        exp_qs = _apply_expense_search(exp_qs, f.q)

    income_rows = list(
        inc_qs.values(
            'id',
            'payment_date',
            'amount',
            'reference',
            'notes',
            'category__name',
            'client__business_name',
            'project_id',
            'created_by__username',
            'created_by__first_name',
            'created_by__last_name',
        )
    )
    expense_rows = list(
        exp_qs.values(
            'id',
            'expense_date',
            'amount',
            'vendor',
            'notes',
            'category__name',
            'project_id',
            'project__client__business_name',
            'created_by__username',
            'created_by__first_name',
            'created_by__last_name',
        )
    )

    rows: list[dict[str, Any]] = []
    for r in income_rows:
        desc = (r.get('notes') or '').strip() or (r.get('reference') or '').strip() or 'Income'
        rows.append({
            'txn_date': r['payment_date'],
            'txn_type': 'income',
            'txn_id': r['id'],
            'description': desc[:200],
            'client': r.get('client__business_name') or '—',
            'project_id': r.get('project_id'),
            'category': r.get('category__name') or '—',
            'reference': r.get('reference') or '—',
            'credit': _money(r['amount']),
            'debit': Decimal('0'),
            'amount': _money(r['amount']),
            'created_by': _created_by_label(r),
            '_ord': 0,
            '_search_blob': ' '.join([
                desc,
                r.get('reference') or '',
                r.get('client__business_name') or '',
                r.get('category__name') or '',
            ]).lower(),
        })
    for r in expense_rows:
        desc = (r.get('notes') or '').strip() or (r.get('vendor') or '').strip() or 'Expense'
        rows.append({
            'txn_date': r['expense_date'],
            'txn_type': 'expense',
            'txn_id': r['id'],
            'description': desc[:200],
            'client': r.get('project__client__business_name') or '—',
            'project_id': r.get('project_id'),
            'category': r.get('category__name') or '—',
            'reference': r.get('vendor') or '—',
            'credit': Decimal('0'),
            'debit': _money(r['amount']),
            'amount': _money(r['amount']),
            'created_by': _created_by_label(r),
            '_ord': 1,
            '_search_blob': ' '.join([
                desc,
                r.get('vendor') or '',
                r.get('project__client__business_name') or '',
                r.get('category__name') or '',
            ]).lower(),
        })

    rows.sort(key=lambda x: (x['txn_date'], x['_ord'], x['txn_id']))
    return rows


def _attach_running_balance(rows: list[dict], opening: Decimal) -> Decimal:
    bal = opening
    for row in rows:
        bal = (bal + row['credit'] - row['debit']).quantize(Decimal('0.01'))
        row['running_balance'] = bal
    return bal


def _sort_for_display(rows: list[dict], sort: str) -> list[dict]:
    sort = (sort or 'newest').lower()
    if sort == 'oldest':
        return rows
    if sort == 'amount':
        return sorted(rows, key=lambda x: (-x['amount'], x['txn_date'], x['txn_id']))
    return list(reversed(rows))


def build_statement(
    f: StatementFilters,
    *,
    page: int | str = 1,
    per_page: int = STATEMENT_PER_PAGE,
    for_export: bool = False,
) -> dict[str, Any]:
    """
    Build statement context: KPIs + paginated (or full) rows with running balance.
    """
    opening = opening_balance(f)
    income_total, expense_total = period_totals(f)
    net = (income_total - expense_total).quantize(Decimal('0.01'))
    closing = (opening + net).quantize(Decimal('0.01'))

    # Full period rows first so running balance stays chronologically correct
    rows = _fetch_period_rows(f, apply_search=False)
    _attach_running_balance(rows, opening)

    if f.q:
        qlow = f.q.lower()
        rows = [r for r in rows if qlow in r.get('_search_blob', '')]

    display_rows = _sort_for_display(rows, f.sort)

    columns = [
        'Date',
        'Type',
        'Description',
        'Client',
        'Project',
        'Category',
        'Reference',
        'Credit',
        'Debit',
        'Running Balance',
        'Created By',
    ]

    def _table_row(r: dict) -> list:
        return [
            r['txn_date'],
            'Income' if r['txn_type'] == 'income' else 'Expense',
            r['description'],
            r['client'],
            f"#{r['project_id']}" if r['project_id'] else '—',
            r['category'],
            r['reference'],
            r['credit'] if r['credit'] else '',
            r['debit'] if r['debit'] else '',
            r['running_balance'],
            r['created_by'],
        ]

    summary = {
        'Opening Balance': opening,
        'Total Income': income_total,
        'Total Expense': expense_total,
        'Net Change': net,
        'Closing Balance': closing,
    }

    base = {
        'report_type': 'statement',
        'report_title': 'Finance Statement',
        'period_label': f.label,
        'filter_start': f.start,
        'filter_end': f.end,
        'opening_balance': opening,
        'income_total': income_total,
        'expense_total': expense_total,
        'net_change': net,
        'closing_balance': closing,
        'summary': summary,
        'columns': columns,
        'filters': f,
        'row_count': len(display_rows),
    }

    if for_export:
        capped = display_rows[:STATEMENT_EXPORT_ROW_CAP]
        base['rows'] = [_table_row(r) for r in capped]
        base['export_rows'] = capped
        if len(display_rows) > STATEMENT_EXPORT_ROW_CAP:
            base['detail_note'] = (
                f'Export limited to first {STATEMENT_EXPORT_ROW_CAP:,} of '
                f'{len(display_rows):,} rows.'
            )
        return base

    paginator = Paginator(display_rows, per_page)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    base['page_obj'] = page_obj
    base['rows'] = [_table_row(r) for r in page_obj.object_list]
    return base


def _cell(value) -> str:
    if value is None or value == '':
        return ''
    if isinstance(value, Decimal):
        return f'{value:.2f}'
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def export_statement_csv(ctx: dict) -> HttpResponse:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([ctx['report_title'], ctx['period_label'],
                f"{ctx['filter_start']} – {ctx['filter_end']}"])
    w.writerow([])
    for k, v in (ctx.get('summary') or {}).items():
        w.writerow([k, _cell(v)])
    w.writerow([])
    w.writerow(ctx.get('columns') or [])
    for row in ctx.get('rows') or []:
        w.writerow([_cell(c) for c in row])
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    fname = f"statement_{ctx['filter_start']}_{ctx['filter_end']}.csv"
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_statement_xlsx(ctx: dict) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Statement'
    ws.append([ctx['report_title']])
    ws.append([f"{ctx['period_label']}: {ctx['filter_start']} – {ctx['filter_end']}"])
    ws.append([])
    for k, v in (ctx.get('summary') or {}).items():
        ws.append([k, _cell(v)])
    ws.append([])
    ws.append(ctx.get('columns') or [])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for row in ctx.get('rows') or []:
        ws.append([_cell(c) for c in row])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"statement_{ctx['filter_start']}_{ctx['filter_end']}.xlsx"
    resp = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_statement_pdf(ctx: dict) -> HttpResponse:
    """Professional bank-statement style PDF (landscape)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'StmtTitle',
        parent=styles['Title'],
        fontSize=16,
        spaceAfter=4,
        textColor=colors.HexColor('#0f172a'),
    )
    meta_style = ParagraphStyle(
        'StmtMeta',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=8,
    )
    story = [
        Paragraph('BThinkX · Finance Statement', title_style),
        Paragraph(
            f"{ctx['period_label']} &nbsp;|&nbsp; "
            f"{ctx['filter_start']} to {ctx['filter_end']}",
            meta_style,
        ),
    ]

    kpi = [
        ['Opening Balance', 'Total Income', 'Total Expense', 'Net Change', 'Closing Balance'],
        [
            f"Rs. {_cell(ctx['opening_balance'])}",
            f"Rs. {_cell(ctx['income_total'])}",
            f"Rs. {_cell(ctx['expense_total'])}",
            f"Rs. {_cell(ctx['net_change'])}",
            f"Rs. {_cell(ctx['closing_balance'])}",
        ],
    ]
    kpi_table = Table(kpi, colWidths=[50 * mm] * 5)
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 10))

    cols = ctx.get('columns') or []
    data = [cols]
    for row in (ctx.get('rows') or [])[:500]:
        data.append([_cell(c) for c in row])

    if len(data) > 1:
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#f8fafc')]),
            ('ALIGN', (7, 1), (9, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(table)
    else:
        story.append(Paragraph('No transactions in this period.', styles['Normal']))

    if ctx.get('detail_note'):
        story.append(Spacer(1, 6))
        story.append(Paragraph(ctx['detail_note'], meta_style))

    story.append(Spacer(1, 8))
    story.append(Paragraph(
        'Generated for management use · Running balance = Opening + Income − Expense',
        meta_style,
    ))

    doc.build(story)
    buf.seek(0)
    fname = f"statement_{ctx['filter_start']}_{ctx['filter_end']}.pdf"
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
