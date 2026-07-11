"""
Founder withdrawals & internal fund management (Phase 8).

All balances are derived from transactions:
  IncomeAllocation / FounderShareAllocation (in)
  − Expense.funding_bucket / FundUsage / FounderWithdrawal / FundTransfer out
  + FundTransfer in

No stored running totals.
"""

from __future__ import annotations

import csv
import io
import json
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal
from typing import Any

from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone

from ..models import (
    AuditEntry,
    Expense,
    Founder,
    FounderShareAllocation,
    FounderWithdrawal,
    FundTransfer,
    FundUsage,
    IncomeAllocation,
    RevenueAllocationBucket,
)
from . import audit as audit_service
from .revenue_allocation import (
    TWOPLACES,
    _money,
    ensure_default_buckets,
    split_amount,
)

DEFAULT_FOUNDERS = (
    ('Achu', Decimal('50.00'), 10),
    ('Veena', Decimal('50.00'), 20),
)

BUCKET_CODE_BY_NAME = {
    'Founder Pool': 'founder_pool',
    'Sales Commission': 'sales_commission',
    'Operations': 'operations',
    'Company Savings': 'company_savings',
}


def ensure_bucket_codes():
    ensure_default_buckets()
    for name, code in BUCKET_CODE_BY_NAME.items():
        RevenueAllocationBucket.objects.filter(name=name, code='').update(code=code)


def ensure_default_founders() -> list[Founder]:
    if Founder.objects.exists():
        return list(Founder.objects.order_by('display_order', 'name'))
    created = []
    for name, pct, order in DEFAULT_FOUNDERS:
        created.append(
            Founder.objects.create(
                name=name,
                percentage=pct,
                active=True,
                display_order=order,
            )
        )
    return created


def active_founders() -> list[Founder]:
    ensure_default_founders()
    return list(Founder.objects.filter(active=True).order_by('display_order', 'name'))


def founder_percentage_total() -> Decimal:
    return _money(
        Founder.objects.filter(active=True).aggregate(t=Sum('percentage'))['t']
    )


def validate_founder_percentages() -> tuple[bool, Decimal]:
    total = founder_percentage_total()
    return total == Decimal('100.00'), total


def get_founder_pool_bucket() -> RevenueAllocationBucket | None:
    ensure_bucket_codes()
    return (
        RevenueAllocationBucket.objects.filter(code='founder_pool').first()
        or RevenueAllocationBucket.objects.filter(name='Founder Pool').first()
    )


def sync_founder_shares_for_income(income, *, actor=None, ip_address: str | None = None):
    """
    Split Founder Pool IncomeAllocation across active founders.
    Idempotent: replaces founder share rows for this income.
    """
    ensure_default_founders()
    pool = get_founder_pool_bucket()
    FounderShareAllocation.objects.filter(income=income).delete()
    if not pool:
        return []

    pool_line = (
        IncomeAllocation.objects.filter(income=income, bucket=pool).first()
    )
    if not pool_line or pool_line.amount <= 0:
        return []

    founders = active_founders()
    if not founders:
        return []

    class _B:
        def __init__(self, f):
            self._f = f
            self.percentage = f.percentage
            self.name = f.name

    fake = [_B(f) for f in founders]
    splits = split_amount(_money(pool_line.amount), fake)

    created = []
    for fake_b, pct, amount in splits:
        founder = next(f for f in founders if f.name == fake_b.name)
        created.append(
            FounderShareAllocation.objects.create(
                income=income,
                income_allocation=pool_line,
                founder=founder,
                percentage=pct,
                amount=amount,
            )
        )

    if created:
        audit_service.log_event(
            category=AuditEntry.EventCategory.FINANCE,
            action='founder_shares_synced',
            object_type='Income',
            object_id=income.pk,
            object_repr=str(income)[:200],
            actor=actor,
            project=income.project,
            after_state={
                'pool_amount': str(pool_line.amount),
                'shares': [
                    {'founder': c.founder.name, 'amount': str(c.amount), 'pct': str(c.percentage)}
                    for c in created
                ],
            },
            ip_address=ip_address,
            note=f'Split Founder Pool Rs. {pool_line.amount} across {len(created)} founders',
        )
    return created


# ── Balance helpers (lifetime or date-bounded) ────────────────────────────────

@dataclass
class DateRange:
    start: date | None = None
    end: date | None = None


def _in_range(field: str, dr: DateRange):
    flt = {}
    if dr.start:
        flt[f'{field}__gte'] = dr.start
    if dr.end:
        flt[f'{field}__lte'] = dr.end
    return flt


def bucket_allocated(bucket: RevenueAllocationBucket, dr: DateRange | None = None) -> Decimal:
    dr = dr or DateRange()
    qs = IncomeAllocation.objects.filter(bucket=bucket)
    if dr.start or dr.end:
        qs = qs.filter(**_in_range('income__payment_date', dr))
    return _money(qs.aggregate(t=Sum('amount'))['t'])


def bucket_expense_out(bucket: RevenueAllocationBucket, dr: DateRange | None = None) -> Decimal:
    dr = dr or DateRange()
    qs = Expense.objects.filter(funding_bucket=bucket)
    if dr.start or dr.end:
        qs = qs.filter(**_in_range('expense_date', dr))
    return _money(qs.aggregate(t=Sum('amount'))['t'])


def bucket_usage_out(bucket: RevenueAllocationBucket, dr: DateRange | None = None) -> Decimal:
    dr = dr or DateRange()
    qs = FundUsage.objects.filter(bucket=bucket)
    if dr.start or dr.end:
        qs = qs.filter(**_in_range('usage_date', dr))
    return _money(qs.aggregate(t=Sum('amount'))['t'])


def bucket_transfer_out(bucket: RevenueAllocationBucket, dr: DateRange | None = None) -> Decimal:
    dr = dr or DateRange()
    qs = FundTransfer.objects.filter(from_bucket=bucket)
    if dr.start or dr.end:
        qs = qs.filter(**_in_range('transfer_date', dr))
    return _money(qs.aggregate(t=Sum('amount'))['t'])


def bucket_transfer_in(bucket: RevenueAllocationBucket, dr: DateRange | None = None) -> Decimal:
    dr = dr or DateRange()
    qs = FundTransfer.objects.filter(to_bucket=bucket)
    if dr.start or dr.end:
        qs = qs.filter(**_in_range('transfer_date', dr))
    return _money(qs.aggregate(t=Sum('amount'))['t'])


def bucket_founder_withdrawals(bucket: RevenueAllocationBucket, dr: DateRange | None = None) -> Decimal:
    """Founder withdrawals only count against the Founder Pool bucket."""
    pool = get_founder_pool_bucket()
    if not pool or bucket.pk != pool.pk:
        return Decimal('0.00')
    dr = dr or DateRange()
    qs = FounderWithdrawal.objects.all()
    if dr.start or dr.end:
        qs = qs.filter(**_in_range('withdrawal_date', dr))
    return _money(qs.aggregate(t=Sum('amount'))['t'])


def bucket_balance(bucket: RevenueAllocationBucket, *, as_of: date | None = None) -> dict[str, Decimal]:
    """
    Lifetime (or as-of) fund balance from all transactions.
    as_of: include only txns on or before this date.
    """
    dr = DateRange(end=as_of) if as_of else DateRange()
    allocated = bucket_allocated(bucket, dr)
    expense_out = bucket_expense_out(bucket, dr)
    usage_out = bucket_usage_out(bucket, dr)
    xfer_out = bucket_transfer_out(bucket, dr)
    xfer_in = bucket_transfer_in(bucket, dr)
    founder_out = bucket_founder_withdrawals(bucket, dr)
    withdrawn = (expense_out + usage_out + xfer_out + founder_out).quantize(TWOPLACES)
    remaining = (allocated + xfer_in - withdrawn).quantize(TWOPLACES)
    return {
        'allocated': allocated,
        'expense_out': expense_out,
        'usage_out': usage_out,
        'transfer_out': xfer_out,
        'transfer_in': xfer_in,
        'founder_withdrawals': founder_out,
        'withdrawn': withdrawn,
        'remaining': remaining,
    }


def all_fund_balances(*, as_of: date | None = None) -> list[dict[str, Any]]:
    ensure_bucket_codes()
    rows = []
    for b in RevenueAllocationBucket.objects.filter(active=True).order_by(
        'display_order', 'name'
    ):
        bal = bucket_balance(b, as_of=as_of)
        rows.append({'bucket': b, **bal})
    return rows


def founder_allocated(founder: Founder, dr: DateRange | None = None) -> Decimal:
    dr = dr or DateRange()
    qs = FounderShareAllocation.objects.filter(founder=founder)
    if dr.start or dr.end:
        qs = qs.filter(**_in_range('income__payment_date', dr))
    return _money(qs.aggregate(t=Sum('amount'))['t'])


def founder_withdrawn(founder: Founder, dr: DateRange | None = None) -> Decimal:
    dr = dr or DateRange()
    qs = FounderWithdrawal.objects.filter(founder=founder)
    if dr.start or dr.end:
        qs = qs.filter(**_in_range('withdrawal_date', dr))
    return _money(qs.aggregate(t=Sum('amount'))['t'])


def founder_balance(founder: Founder, *, as_of: date | None = None) -> dict[str, Decimal]:
    dr = DateRange(end=as_of) if as_of else DateRange()
    allocated = founder_allocated(founder, dr)
    withdrawn = founder_withdrawn(founder, dr)
    return {
        'allocated': allocated,
        'withdrawn': withdrawn,
        'remaining': (allocated - withdrawn).quantize(TWOPLACES),
    }


def available_bucket_balance(
    bucket: RevenueAllocationBucket,
    *,
    exclude_expense: Expense | None = None,
) -> Decimal:
    """Live remaining balance; optionally ignore an expense being edited."""
    remaining = bucket_balance(bucket)['remaining']
    if (
        exclude_expense
        and exclude_expense.pk
        and exclude_expense.funding_bucket_id == bucket.pk
    ):
        remaining = (remaining + _money(exclude_expense.amount)).quantize(TWOPLACES)
    return remaining


def backfill_founder_shares(*, limit: int = 2000, actor=None) -> int:
    """Create missing founder share rows for incomes that already have Founder Pool allocations."""
    ensure_default_founders()
    pool = get_founder_pool_bucket()
    if not pool:
        return 0
    income_ids = (
        IncomeAllocation.objects.filter(bucket=pool)
        .exclude(
            income_id__in=FounderShareAllocation.objects.values_list(
                'income_id', flat=True
            )
        )
        .values_list('income_id', flat=True)
        .distinct()[:limit]
    )
    from ..models import Income

    n = 0
    for income in Income.objects.filter(pk__in=list(income_ids)).iterator(chunk_size=100):
        sync_founder_shares_for_income(income, actor=actor)
        n += 1
    return n


# ── Mutations ────────────────────────────────────────────────────────────────

def create_founder_withdrawal(
    *,
    actor,
    founder: Founder,
    amount,
    withdrawal_date=None,
    reference: str = '',
    notes: str = '',
    ip_address: str | None = None,
) -> FounderWithdrawal:
    amount = _money(amount)
    if amount <= 0:
        raise ValidationError('Withdrawal amount must be greater than zero.')
    bal = founder_balance(founder)
    if amount > bal['remaining']:
        raise ValidationError(
            f'Insufficient balance for {founder.name}. '
            f'Available Rs. {bal["remaining"]:.2f}, requested Rs. {amount:.2f}.'
        )
    pool = get_founder_pool_bucket()
    if pool:
        pool_rem = bucket_balance(pool)['remaining']
        if amount > pool_rem:
            raise ValidationError(
                f'Insufficient Founder Pool balance. '
                f'Available Rs. {pool_rem:.2f}, requested Rs. {amount:.2f}.'
            )
    w = FounderWithdrawal.objects.create(
        founder=founder,
        amount=amount,
        withdrawal_date=withdrawal_date or timezone.localdate(),
        reference=(reference or '')[:120],
        notes=notes or '',
        created_by=actor,
    )
    audit_service.log_event(
        category=AuditEntry.EventCategory.FINANCE,
        action='founder_withdrawal_created',
        object_type='FounderWithdrawal',
        object_id=w.pk,
        object_repr=str(w)[:200],
        actor=actor,
        after_state={
            'founder': founder.name,
            'amount': str(amount),
            'date': str(w.withdrawal_date),
            'reference': w.reference,
            'remaining_after': str(founder_balance(founder)['remaining']),
        },
        ip_address=ip_address,
    )
    return w


def create_fund_transfer(
    *,
    actor,
    from_bucket: RevenueAllocationBucket,
    to_bucket: RevenueAllocationBucket,
    amount,
    transfer_date=None,
    reason: str = '',
    ip_address: str | None = None,
) -> FundTransfer:
    amount = _money(amount)
    if amount <= 0:
        raise ValidationError('Transfer amount must be greater than zero.')
    if from_bucket.pk == to_bucket.pk:
        raise ValidationError('Cannot transfer a fund to itself.')
    if from_bucket.code == 'founder_pool' or from_bucket.name == 'Founder Pool':
        raise ValidationError(
            'Cannot transfer out of Founder Pool. Use Founder Withdraw for payouts.'
        )
    bal = bucket_balance(from_bucket)
    if amount > bal['remaining']:
        raise ValidationError(
            f'Insufficient balance in {from_bucket.name}. '
            f'Available Rs. {bal["remaining"]:.2f}, requested Rs. {amount:.2f}.'
        )
    t = FundTransfer.objects.create(
        from_bucket=from_bucket,
        to_bucket=to_bucket,
        amount=amount,
        transfer_date=transfer_date or timezone.localdate(),
        reason=(reason or '')[:255],
        created_by=actor,
    )
    audit_service.log_event(
        category=AuditEntry.EventCategory.FINANCE,
        action='fund_transfer_created',
        object_type='FundTransfer',
        object_id=t.pk,
        object_repr=str(t)[:200],
        actor=actor,
        after_state={
            'from': from_bucket.name,
            'to': to_bucket.name,
            'amount': str(amount),
            'date': str(t.transfer_date),
            'reason': t.reason,
        },
        ip_address=ip_address,
    )
    return t


# ── Dashboards & reports ─────────────────────────────────────────────────────

@dataclass
class FundFilter:
    date_from: date | None = None
    date_to: date | None = None
    period: str = 'month'
    label: str = 'This Month'


def resolve_fund_period(period, *, date_from=None, date_to=None, ref=None):
    today = ref or timezone.localdate()
    period = (period or 'month').strip().lower()
    if period == 'today':
        return today, today, 'Today'
    if period == 'week':
        start = today - timedelta(days=today.weekday())
        return start, start + timedelta(days=6), 'This Week'
    if period == 'year':
        return date(today.year, 1, 1), date(today.year, 12, 31), 'This Year'
    if period == 'all':
        return None, None, 'All time'
    if period == 'custom' and date_from and date_to:
        if date_from > date_to:
            date_from, date_to = date_to, date_from
        return date_from, date_to, 'Custom'
    start = today.replace(day=1)
    end = today.replace(day=monthrange(today.year, today.month)[1])
    return start, end, 'This Month'


def filters_from_cleaned(data: dict) -> FundFilter:
    period = data.get('period') or 'month'
    start, end, label = resolve_fund_period(
        period, date_from=data.get('date_from'), date_to=data.get('date_to')
    )
    return FundFilter(date_from=start, date_to=end, period=period, label=label)


def get_founder_dashboard(f: FundFilter | None = None) -> dict[str, Any]:
    ensure_default_founders()
    ensure_bucket_codes()
    f = f or FundFilter()
    # Lifetime cards for available balance; period for activity tables
    lifetime = DateRange()
    period = DateRange(start=f.date_from, end=f.date_to)

    cards = []
    for founder in active_founders():
        life = founder_balance(founder)
        period_alloc = founder_allocated(founder, period) if (f.date_from or f.date_to) else life['allocated']
        period_wd = founder_withdrawn(founder, period) if (f.date_from or f.date_to) else life['withdrawn']
        cards.append({
            'founder': founder,
            'allocated': life['allocated'],
            'withdrawn': life['withdrawn'],
            'remaining': life['remaining'],
            'period_allocated': period_alloc,
            'period_withdrawn': period_wd,
            'pending': life['remaining'],
        })

    withdrawals = list(
        FounderWithdrawal.objects.select_related('founder', 'created_by')
        .order_by('-withdrawal_date', '-id')[:40]
    )
    if f.date_from or f.date_to:
        wqs = FounderWithdrawal.objects.select_related('founder', 'created_by')
        if f.date_from:
            wqs = wqs.filter(withdrawal_date__gte=f.date_from)
        if f.date_to:
            wqs = wqs.filter(withdrawal_date__lte=f.date_to)
        withdrawals = list(wqs.order_by('-withdrawal_date', '-id')[:40])

    recent_shares = list(
        FounderShareAllocation.objects.select_related(
            'founder', 'income', 'income__client'
        ).order_by('-income__payment_date', '-id')[:30]
    )

    total_remaining = sum((c['remaining'] for c in cards), Decimal('0'))
    total_allocated = sum((c['allocated'] for c in cards), Decimal('0'))
    total_withdrawn = sum((c['withdrawn'] for c in cards), Decimal('0'))

    return {
        'cards': cards,
        'withdrawals': withdrawals,
        'recent_shares': recent_shares,
        'total_allocated': total_allocated,
        'total_withdrawn': total_withdrawn,
        'total_remaining': total_remaining,
        'period_label': f.label,
        'filter_start': f.date_from,
        'filter_end': f.date_to,
        'founders_ok': validate_founder_percentages()[0],
        'founders_pct_total': validate_founder_percentages()[1],
        'chart_labels': json.dumps([c['founder'].name for c in cards]),
        'chart_allocated': json.dumps([float(c['allocated']) for c in cards]),
        'chart_remaining': json.dumps([float(c['remaining']) for c in cards]),
    }


def get_fund_balances_report() -> dict[str, Any]:
    rows = all_fund_balances()
    return {
        'report_title': 'Current Fund Balances',
        'report_type': 'fund_balances',
        'period_label': 'As of today',
        'filter_start': timezone.localdate(),
        'filter_end': timezone.localdate(),
        'columns': [
            'Fund', 'Allocated', 'Expenses', 'Usage', 'Transfers Out',
            'Transfers In', 'Founder Withdrawals', 'Remaining',
        ],
        'rows': [
            [
                r['bucket'].name,
                r['allocated'],
                r['expense_out'],
                r['usage_out'],
                r['transfer_out'],
                r['transfer_in'],
                r['founder_withdrawals'],
                r['remaining'],
            ]
            for r in rows
        ],
        'summary': {
            'Funds': len(rows),
            'Total Remaining': sum((r['remaining'] for r in rows), Decimal('0')),
        },
        'balance_rows': rows,
    }


def get_fund_utilization_report(f: FundFilter) -> dict[str, Any]:
    dr = DateRange(start=f.date_from, end=f.date_to)
    rows = []
    for b in RevenueAllocationBucket.objects.filter(active=True).order_by(
        'display_order', 'name'
    ):
        allocated = bucket_allocated(b, dr)
        used = (
            bucket_expense_out(b, dr)
            + bucket_usage_out(b, dr)
            + bucket_founder_withdrawals(b, dr)
        )
        xfer_net = bucket_transfer_in(b, dr) - bucket_transfer_out(b, dr)
        rows.append({
            'bucket': b,
            'allocated': allocated,
            'used': used,
            'transfer_net': xfer_net,
            'utilization_pct': float((used / allocated) * 100) if allocated else 0.0,
        })
    return {
        'report_title': 'Fund Utilization Report',
        'report_type': 'fund_utilization',
        'period_label': f.label,
        'filter_start': f.date_from or date(2000, 1, 1),
        'filter_end': f.date_to or timezone.localdate(),
        'columns': ['Fund', 'Allocated', 'Used', 'Transfer Net', 'Utilization %'],
        'rows': [
            [r['bucket'].name, r['allocated'], r['used'], r['transfer_net'],
             f"{r['utilization_pct']:.1f}%"]
            for r in rows
        ],
        'summary': {
            'Allocated': sum((r['allocated'] for r in rows), Decimal('0')),
            'Used': sum((r['used'] for r in rows), Decimal('0')),
        },
        'util_rows': rows,
    }


def get_transfer_report(f: FundFilter) -> dict[str, Any]:
    qs = FundTransfer.objects.select_related(
        'from_bucket', 'to_bucket', 'created_by'
    ).order_by('-transfer_date', '-id')
    if f.date_from:
        qs = qs.filter(transfer_date__gte=f.date_from)
    if f.date_to:
        qs = qs.filter(transfer_date__lte=f.date_to)
    items = list(qs[:500])
    return {
        'report_title': 'Fund Transfer Report',
        'report_type': 'fund_transfers',
        'period_label': f.label,
        'filter_start': f.date_from or date(2000, 1, 1),
        'filter_end': f.date_to or timezone.localdate(),
        'columns': ['Date', 'From', 'To', 'Amount', 'Reason', 'Created By'],
        'rows': [
            [
                t.transfer_date,
                t.from_bucket.name,
                t.to_bucket.name,
                t.amount,
                t.reason or '—',
                (t.created_by.get_full_name() or t.created_by.get_username())
                if t.created_by_id else '—',
            ]
            for t in items
        ],
        'summary': {
            'Transfers': len(items),
            'Total Amount': sum((t.amount for t in items), Decimal('0')),
        },
        'transfers': items,
    }


def get_founder_statement(founder: Founder, f: FundFilter) -> dict[str, Any]:
    shares = FounderShareAllocation.objects.filter(founder=founder).select_related(
        'income', 'income__client'
    )
    wds = FounderWithdrawal.objects.filter(founder=founder).select_related('created_by')
    if f.date_from:
        shares = shares.filter(income__payment_date__gte=f.date_from)
        wds = wds.filter(withdrawal_date__gte=f.date_from)
    if f.date_to:
        shares = shares.filter(income__payment_date__lte=f.date_to)
        wds = wds.filter(withdrawal_date__lte=f.date_to)

    lines = []
    for s in shares.order_by('income__payment_date', 'id'):
        lines.append({
            'date': s.income.payment_date,
            'type': 'Allocation',
            'detail': f'Income #{s.income_id}',
            'credit': s.amount,
            'debit': Decimal('0'),
        })
    for w in wds.order_by('withdrawal_date', 'id'):
        lines.append({
            'date': w.withdrawal_date,
            'type': 'Withdrawal',
            'detail': w.reference or w.notes or 'Withdrawal',
            'credit': Decimal('0'),
            'debit': w.amount,
        })
    lines.sort(key=lambda x: (x['date'], 0 if x['type'] == 'Allocation' else 1))

    # Opening = lifetime before range
    opening = Decimal('0.00')
    if f.date_from:
        before = DateRange(end=f.date_from - timedelta(days=1))
        opening = (
            founder_allocated(founder, before) - founder_withdrawn(founder, before)
        ).quantize(TWOPLACES)

    bal = opening
    rows = []
    for line in lines:
        bal = (bal + line['credit'] - line['debit']).quantize(TWOPLACES)
        rows.append([
            line['date'],
            line['type'],
            line['detail'],
            line['credit'] or '',
            line['debit'] or '',
            bal,
        ])

    life = founder_balance(founder)
    return {
        'report_title': f'Founder Statement — {founder.name}',
        'report_type': 'founder_statement',
        'period_label': f.label,
        'filter_start': f.date_from or date(2000, 1, 1),
        'filter_end': f.date_to or timezone.localdate(),
        'founder': founder,
        'opening_balance': opening,
        'closing_balance': life['remaining'] if not f.date_to else bal,
        'columns': ['Date', 'Type', 'Detail', 'Credit', 'Debit', 'Balance'],
        'rows': rows,
        'summary': {
            'Opening': opening,
            'Allocated (period)': founder_allocated(
                founder, DateRange(start=f.date_from, end=f.date_to)
            ),
            'Withdrawn (period)': founder_withdrawn(
                founder, DateRange(start=f.date_from, end=f.date_to)
            ),
            'Available (lifetime)': life['remaining'],
        },
    }


def _cell(value) -> str:
    if value is None or value == '':
        return ''
    if isinstance(value, Decimal):
        return f'{value:.2f}'
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def export_report_csv(ctx: dict) -> HttpResponse:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([ctx['report_title'], ctx.get('period_label', ''),
                f"{ctx.get('filter_start')} – {ctx.get('filter_end')}"])
    w.writerow([])
    for k, v in (ctx.get('summary') or {}).items():
        w.writerow([k, _cell(v)])
    w.writerow([])
    w.writerow(ctx.get('columns') or [])
    for row in ctx.get('rows') or []:
        w.writerow([_cell(c) for c in row])
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    fname = f"{ctx.get('report_type', 'report')}_{ctx.get('filter_start')}_{ctx.get('filter_end')}.csv"
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_report_xlsx(ctx: dict) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (ctx.get('report_title') or 'Report')[:31]
    ws.append([ctx['report_title']])
    ws.append([f"{ctx.get('period_label')}: {ctx.get('filter_start')} – {ctx.get('filter_end')}"])
    ws.append([])
    for k, v in (ctx.get('summary') or {}).items():
        ws.append([k, _cell(v)])
    ws.append([])
    ws.append(ctx.get('columns') or [])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for row in ctx.get('rows') or []:
        ws.append([_cell(c) for c in row])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"{ctx.get('report_type', 'report')}_{ctx.get('filter_start')}_{ctx.get('filter_end')}.xlsx"
    resp = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_report_pdf(ctx: dict) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(ctx['report_title'], styles['Title']),
        Paragraph(
            f"{ctx.get('period_label')} · {ctx.get('filter_start')} – {ctx.get('filter_end')}",
            styles['Normal'],
        ),
        Spacer(1, 8),
    ]
    for k, v in (ctx.get('summary') or {}).items():
        story.append(Paragraph(f'<b>{k}:</b> {_cell(v)}', styles['Normal']))
    story.append(Spacer(1, 8))
    data = [ctx.get('columns') or []] + [
        [_cell(c) for c in row] for row in (ctx.get('rows') or [])[:300]
    ]
    if data and data[0]:
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        story.append(table)
    doc.build(story)
    buf.seek(0)
    fname = f"{ctx.get('report_type', 'report')}_{ctx.get('filter_start')}_{ctx.get('filter_end')}.pdf"
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
