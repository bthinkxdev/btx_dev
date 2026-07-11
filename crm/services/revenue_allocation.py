"""
Configurable revenue allocation — split each Income into management funds.

Not accounting: no journals/ledgers. Percentages come from RevenueAllocationBucket
settings (must total 100% when active). Allocations are idempotent per income.
"""

from __future__ import annotations

import csv
import io
import json
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import ROUND_HALF_UP, Decimal
from typing import Any

from django.db import transaction
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.utils import timezone

from ..models import (
    AuditEntry,
    FundUsage,
    Income,
    IncomeAllocation,
    RevenueAllocationBucket,
)
from . import audit as audit_service

DEFAULT_BUCKETS = (
    # name, code, percentage, color, order, usage_label
    ('Founder Pool', 'founder_pool', Decimal('20.00'), '#4f46e5', 10, 'used'),
    ('Sales Commission', 'sales_commission', Decimal('10.00'), '#10b981', 20, 'paid'),
    ('Operations', 'operations', Decimal('40.00'), '#f59e0b', 30, 'spent'),
    ('Company Savings', 'company_savings', Decimal('30.00'), '#0ea5e9', 40, 'used'),
)

TWOPLACES = Decimal('0.01')


def _money(value) -> Decimal:
    if value is None:
        return Decimal('0')
    return Decimal(value).quantize(TWOPLACES, rounding=ROUND_HALF_UP)


def ensure_default_buckets() -> list[RevenueAllocationBucket]:
    """Seed default allocation buckets if none exist; backfill codes on known names."""
    if not RevenueAllocationBucket.objects.exists():
        created = []
        for name, code, pct, color, order, usage in DEFAULT_BUCKETS:
            created.append(
                RevenueAllocationBucket.objects.create(
                    name=name,
                    code=code,
                    percentage=pct,
                    color=color,
                    display_order=order,
                    active=True,
                    usage_label=usage,
                )
            )
        return created
    for name, code, *_rest in DEFAULT_BUCKETS:
        RevenueAllocationBucket.objects.filter(name=name, code='').update(code=code)
    return list(RevenueAllocationBucket.objects.order_by('display_order', 'name'))


def active_buckets() -> list[RevenueAllocationBucket]:
    ensure_default_buckets()
    return list(
        RevenueAllocationBucket.objects.filter(active=True).order_by(
            'display_order', 'name'
        )
    )


def active_percentage_total() -> Decimal:
    total = (
        RevenueAllocationBucket.objects.filter(active=True).aggregate(
            t=Sum('percentage')
        )['t']
    )
    return _money(total)


def validate_active_percentages() -> tuple[bool, Decimal]:
    total = active_percentage_total()
    return total == Decimal('100.00'), total


def split_amount(total: Decimal, buckets: list[RevenueAllocationBucket]) -> list[tuple[RevenueAllocationBucket, Decimal, Decimal]]:
    """
    Split total across buckets by percentage. Uses largest-remainder so
    line amounts sum exactly to total (no drift from rounding).
    Returns list of (bucket, percentage, amount).
    """
    total = _money(total)
    if not buckets or total == 0:
        return [(b, _money(b.percentage), Decimal('0.00')) for b in buckets]

    pct_sum = sum((_money(b.percentage) for b in buckets), Decimal('0'))
    if pct_sum <= 0:
        return [(b, _money(b.percentage), Decimal('0.00')) for b in buckets]

    # Scale if somehow not 100 (should be prevented by settings UI)
    scale = Decimal('100.00') / pct_sum if pct_sum != Decimal('100.00') else Decimal('1')

    raw: list[tuple[RevenueAllocationBucket, Decimal, Decimal, Decimal]] = []
    # (bucket, pct, floored_amount, remainder_fraction)
    for b in buckets:
        pct = _money(b.percentage) * scale
        exact = (total * pct / Decimal('100')).quantize(Decimal('0.0001'))
        floored = exact.quantize(TWOPLACES, rounding=ROUND_HALF_UP)
        # Use remainder for largest-remainder correction
        remainder = exact - floored
        raw.append((b, pct, floored, remainder))

    assigned = sum((r[2] for r in raw), Decimal('0'))
    diff_cents = int(((total - assigned) / TWOPLACES).to_integral_value())
    # Sort by remainder descending to distribute leftover/missing paise
    ordered = sorted(enumerate(raw), key=lambda ir: ir[1][3], reverse=(diff_cents > 0))
    amounts = [r[2] for r in raw]
    step = TWOPLACES if diff_cents > 0 else -TWOPLACES
    for i in range(abs(diff_cents)):
        idx = ordered[i % len(ordered)][0]
        amounts[idx] = (amounts[idx] + step).quantize(TWOPLACES)

    return [(raw[i][0], raw[i][1], amounts[i]) for i in range(len(raw))]


def sync_income_allocations(
    income: Income,
    *,
    actor=None,
    ip_address: str | None = None,
    note: str = '',
) -> list[IncomeAllocation]:
    """
    Idempotent: replace allocation lines for this income from active buckets.
    No duplicate rows — UniqueConstraint (income, bucket).
    """
    buckets = active_buckets()
    ok, total_pct = validate_active_percentages()
    if not buckets:
        IncomeAllocation.objects.filter(income=income).delete()
        return []
    if not ok:
        # Still allocate using scaled percentages so income is never left unsplit
        pass

    splits = split_amount(_money(income.amount), buckets)
    before = [
        {'bucket_id': a.bucket_id, 'amount': str(a.amount), 'percentage': str(a.percentage)}
        for a in IncomeAllocation.objects.filter(income=income).select_related('bucket')
    ]

    with transaction.atomic():
        IncomeAllocation.objects.filter(income=income).delete()
        created: list[IncomeAllocation] = []
        for bucket, pct, amount in splits:
            created.append(
                IncomeAllocation.objects.create(
                    income=income,
                    bucket=bucket,
                    percentage=pct,
                    amount=amount,
                )
            )
        from crm.services.fund_management import sync_founder_shares_for_income

        sync_founder_shares_for_income(income)

    after = [
        {'bucket_id': a.bucket_id, 'bucket': a.bucket.name,
         'amount': str(a.amount), 'percentage': str(a.percentage)}
        for a in created
    ]
    audit_service.log_event(
        category=AuditEntry.EventCategory.FINANCE,
        action='income_allocation_synced',
        object_type='Income',
        object_id=income.pk,
        object_repr=str(income)[:200],
        actor=actor,
        project=income.project,
        before_state={'allocations': before} if before else None,
        after_state={
            'income_amount': str(income.amount),
            'allocations': after,
            'active_pct_total': str(total_pct),
        },
        ip_address=ip_address,
        note=note or f'Allocated Rs. {income.amount} across {len(created)} funds',
    )
    return created


def clear_income_allocations(
    income_id: int,
    *,
    actor=None,
    project=None,
    income_repr: str = '',
    before_state: dict | None = None,
    ip_address: str | None = None,
):
    """Called around income delete (CASCADE also removes rows)."""
    audit_service.log_event(
        category=AuditEntry.EventCategory.FINANCE,
        action='income_allocation_cleared',
        object_type='Income',
        object_id=income_id,
        object_repr=(income_repr or f'Income #{income_id}')[:200],
        actor=actor,
        project=project,
        before_state=before_state,
        ip_address=ip_address,
        note='Allocations removed with income',
    )


def backfill_missing_allocations(*, actor=None, limit: int = 5000) -> int:
    """Allocate for incomes that have no allocation rows yet."""
    ensure_default_buckets()
    missing = (
        Income.objects.annotate(n=Count('allocations'))
        .filter(n=0)
        .order_by('id')[:limit]
    )
    count = 0
    for income in missing:
        sync_income_allocations(income, actor=actor, note='Backfill missing allocation')
        count += 1
    return count


@dataclass
class AllocationFilters:
    date_from: date | None = None
    date_to: date | None = None
    project_id: int | None = None
    client_id: int | None = None
    category_id: int | None = None  # income category
    period: str = 'month'
    label: str = 'This Month'


def resolve_alloc_period(
    period: str,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    ref: date | None = None,
) -> tuple[date, date, str]:
    today = ref or timezone.localdate()
    period = (period or 'month').strip().lower()
    if period == 'today':
        return today, today, 'Today'
    if period == 'week':
        start = today - timedelta(days=today.weekday())
        return start, start + timedelta(days=6), 'This Week'
    if period == 'year':
        return date(today.year, 1, 1), date(today.year, 12, 31), 'This Year'
    if period == 'custom' and date_from and date_to:
        if date_from > date_to:
            date_from, date_to = date_to, date_from
        return date_from, date_to, 'Custom'
    start = today.replace(day=1)
    end = today.replace(day=monthrange(today.year, today.month)[1])
    return start, end, 'This Month'


def filters_from_cleaned(data: dict) -> AllocationFilters:
    period = data.get('period') or 'month'
    start, end, label = resolve_alloc_period(
        period,
        date_from=data.get('date_from'),
        date_to=data.get('date_to'),
    )
    project = data.get('project')
    client = data.get('client')
    category = data.get('category')
    return AllocationFilters(
        date_from=start,
        date_to=end,
        project_id=getattr(project, 'pk', project) or None,
        client_id=getattr(client, 'pk', client) or None,
        category_id=getattr(category, 'pk', category) or None,
        period=period,
        label=label,
    )


def _income_qs(f: AllocationFilters):
    qs = Income.objects.filter(
        payment_date__gte=f.date_from,
        payment_date__lte=f.date_to,
    )
    if f.project_id:
        qs = qs.filter(project_id=f.project_id)
    if f.client_id:
        qs = qs.filter(client_id=f.client_id)
    if f.category_id:
        qs = qs.filter(category_id=f.category_id)
    return qs


def _allocation_qs(f: AllocationFilters):
    return IncomeAllocation.objects.filter(
        income__payment_date__gte=f.date_from,
        income__payment_date__lte=f.date_to,
        **({} if not f.project_id else {'income__project_id': f.project_id}),
        **({} if not f.client_id else {'income__client_id': f.client_id}),
        **({} if not f.category_id else {'income__category_id': f.category_id}),
    ).select_related('bucket', 'income', 'income__client', 'income__category')


def _usage_qs(f: AllocationFilters):
    qs = FundUsage.objects.filter(
        usage_date__gte=f.date_from,
        usage_date__lte=f.date_to,
    )
    if f.project_id:
        qs = qs.filter(project_id=f.project_id)
    return qs.select_related('bucket', 'project', 'created_by')


def get_allocation_dashboard(f: AllocationFilters) -> dict[str, Any]:
    ensure_default_buckets()
    backfill_missing_allocations(limit=200)

    buckets = list(
        RevenueAllocationBucket.objects.filter(active=True).order_by(
            'display_order', 'name'
        )
    )
    alloc_agg = {
        r['bucket_id']: _money(r['total'])
        for r in (
            _allocation_qs(f)
            .values('bucket_id')
            .annotate(total=Sum('amount'))
        )
    }

    from crm.services.fund_management import (
        DateRange,
        bucket_balance,
        bucket_expense_out,
        bucket_founder_withdrawals,
        bucket_transfer_out,
        bucket_usage_out,
    )

    cards = []
    dist_labels, dist_values, dist_colors = [], [], []
    util_labels, util_alloc, util_used = [], [], []
    dr = DateRange(start=f.date_from, end=f.date_to)
    for b in buckets:
        allocated = alloc_agg.get(b.pk, Decimal('0'))
        # Align with live remaining formula (period-scoped outs)
        used = (
            bucket_expense_out(b, dr)
            + bucket_usage_out(b, dr)
            + bucket_transfer_out(b, dr)
            + bucket_founder_withdrawals(b, dr)
        )
        life = bucket_balance(b)
        remaining = life['remaining']
        cards.append({
            'bucket': b,
            'allocated': allocated,
            'used': used,
            'remaining': remaining,
            'lifetime_allocated': life['allocated'],
            'usage_label': b.get_usage_label_display(),
            'color': b.color,
            'percentage': b.percentage,
        })
        dist_labels.append(b.name)
        dist_values.append(float(allocated))
        dist_colors.append(b.color or '#64748b')
        util_labels.append(b.name)
        util_alloc.append(float(allocated))
        util_used.append(float(used))

    # Monthly allocation (last 6 months) — by payment_date month via Python group
    today = timezone.localdate()
    months: list[date] = []
    cursor = today.replace(day=1)
    for _ in range(6):
        months.append(cursor)
        cursor = (cursor - timedelta(days=1)).replace(day=1)
    months.reverse()
    range_start = months[0]
    range_end = today.replace(day=monthrange(today.year, today.month)[1])

    monthly_rows = list(
        IncomeAllocation.objects.filter(
            income__payment_date__gte=range_start,
            income__payment_date__lte=range_end,
        )
        .values('income__payment_date', 'amount')
    )
    month_totals: dict[tuple[int, int], Decimal] = {
        (m.year, m.month): Decimal('0') for m in months
    }
    for row in monthly_rows:
        d = row['income__payment_date']
        key = (d.year, d.month)
        if key in month_totals:
            month_totals[key] += _money(row['amount'])

    monthly_labels = [m.strftime('%b %Y') for m in months]
    monthly_values = [float(month_totals[(m.year, m.month)]) for m in months]

    recent_allocations = list(
        _allocation_qs(f)
        .select_related(
            'bucket', 'income', 'income__client', 'income__category', 'income__project'
        )
        .order_by('-income__payment_date', '-id')[:25]
    )
    fund_history = list(
        _allocation_qs(f)
        .select_related('bucket', 'income', 'income__client')
        .order_by('-created_at', '-id')[:40]
    )
    recent_usages = list(_usage_qs(f).order_by('-usage_date', '-id')[:20])

    income_total = _money(
        _income_qs(f).aggregate(t=Sum('amount'))['t']
    )
    allocated_total = sum((c['allocated'] for c in cards), Decimal('0'))
    used_total = sum((c['used'] for c in cards), Decimal('0'))
    remaining_total = sum((c['remaining'] for c in cards), Decimal('0'))
    ok_pct, pct_total = validate_active_percentages()

    return {
        'filter_start': f.date_from,
        'filter_end': f.date_to,
        'period_label': f.label,
        'cards': cards,
        'income_total': income_total,
        'allocated_total': allocated_total,
        'used_total': used_total,
        'remaining_total': remaining_total,
        'settings_ok': ok_pct,
        'settings_pct_total': pct_total,
        'recent_allocations': recent_allocations,
        'fund_history': fund_history,
        'recent_usages': recent_usages,
        'chart_dist_labels': json.dumps(dist_labels),
        'chart_dist_values': json.dumps(dist_values),
        'chart_dist_colors': json.dumps(dist_colors),
        'chart_util_labels': json.dumps(util_labels),
        'chart_util_alloc': json.dumps(util_alloc),
        'chart_util_used': json.dumps(util_used),
        'chart_monthly_labels': json.dumps(monthly_labels),
        'chart_monthly_values': json.dumps(monthly_values),
        'columns': [
            'Date', 'Income #', 'Client', 'Bucket', 'Percentage', 'Amount',
        ],
        'rows': [
            [
                a.income.payment_date,
                a.income_id,
                a.income.client.business_name if a.income.client_id else '—',
                a.bucket.name,
                a.percentage,
                a.amount,
            ]
            for a in recent_allocations
        ],
        'summary': {
            'Income in period': income_total,
            'Allocated': allocated_total,
            'Used': used_total,
            'Remaining (live)': remaining_total,
        },
        'report_type': 'allocation',
        'report_title': 'Revenue Allocation',
    }


def create_fund_usage(*, actor, ip_address: str | None = None, **fields) -> FundUsage:
    from django.core.exceptions import ValidationError

    raise ValidationError(
        'Record Usage is retired. Create an Expense and set Funding Source '
        'to draw from a fund. Founder Pool payouts use Founder Withdraw.'
    )


def _cell(value) -> str:
    if value is None:
        return ''
    if isinstance(value, Decimal):
        return f'{value:.2f}'
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def export_allocation_csv(ctx: dict) -> HttpResponse:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([ctx['report_title'], ctx['period_label'],
                f"{ctx['filter_start']} – {ctx['filter_end']}"])
    w.writerow([])
    for k, v in (ctx.get('summary') or {}).items():
        w.writerow([k, _cell(v)])
    w.writerow([])
    w.writerow(['Fund', 'Allocated', 'Used', 'Remaining', '% Setting'])
    for c in ctx.get('cards') or []:
        w.writerow([
            c['bucket'].name,
            _cell(c['allocated']),
            _cell(c['used']),
            _cell(c['remaining']),
            _cell(c['percentage']),
        ])
    w.writerow([])
    w.writerow(ctx.get('columns') or [])
    for row in ctx.get('rows') or []:
        w.writerow([_cell(c) for c in row])
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    fname = f"allocation_{ctx['filter_start']}_{ctx['filter_end']}.csv"
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_allocation_xlsx(ctx: dict) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Allocation'
    ws.append([ctx['report_title']])
    ws.append([f"{ctx['period_label']}: {ctx['filter_start']} – {ctx['filter_end']}"])
    ws.append([])
    for k, v in (ctx.get('summary') or {}).items():
        ws.append([k, _cell(v)])
    ws.append([])
    ws.append(['Fund', 'Allocated', 'Used', 'Remaining', '% Setting'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for c in ctx.get('cards') or []:
        ws.append([
            c['bucket'].name,
            _cell(c['allocated']),
            _cell(c['used']),
            _cell(c['remaining']),
            _cell(c['percentage']),
        ])
    ws.append([])
    ws.append(ctx.get('columns') or [])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for row in ctx.get('rows') or []:
        ws.append([_cell(c) for c in row])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"allocation_{ctx['filter_start']}_{ctx['filter_end']}.xlsx"
    resp = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_allocation_pdf(ctx: dict) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph('Revenue Allocation', styles['Title']),
        Paragraph(
            f"{ctx['period_label']} · {ctx['filter_start']} – {ctx['filter_end']}",
            styles['Normal'],
        ),
        Spacer(1, 8),
    ]
    for k, v in (ctx.get('summary') or {}).items():
        story.append(Paragraph(f'<b>{k}:</b> Rs. {_cell(v)}', styles['Normal']))
    story.append(Spacer(1, 8))

    fund_data = [['Fund', 'Allocated', 'Used', 'Remaining', '%']]
    for c in ctx.get('cards') or []:
        fund_data.append([
            c['bucket'].name,
            _cell(c['allocated']),
            _cell(c['used']),
            _cell(c['remaining']),
            _cell(c['percentage']),
        ])
    ft = Table(fund_data, repeatRows=1)
    ft.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    story.append(ft)
    story.append(Spacer(1, 10))

    cols = ctx.get('columns') or []
    data = [cols] + [[_cell(c) for c in row] for row in (ctx.get('rows') or [])[:200]]
    if len(data) > 1:
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
        ]))
        story.append(table)

    doc.build(story)
    buf.seek(0)
    fname = f"allocation_{ctx['filter_start']}_{ctx['filter_end']}.pdf"
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
